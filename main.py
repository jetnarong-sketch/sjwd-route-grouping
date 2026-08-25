# Re-generate the clean main.py file without any trailing execution code
clean_main_code = """from datetime import datetime, date
import io
import os
import json
import time
import re
import openpyxl
import pandas as pd
import streamlit as st

# --- USER CREDENTIALS & ROLES DATABASE ---
USER_DB = {
    "admin": {
        "password": "adminpassword123",
        "name": "Admin Ball",
        "role": "Admin",
    },
    "pm": {
        "password": "pmpassword123",
        "name": "Project Manager",
        "role": "Project Manager",
    },
    "operator": {
        "password": "operatorpassword123",
        "name": "Operator Team",
        "role": "Operator",
    },
}

# --- TRANSLATION DICTIONARY ---
T = {
    "TH": {
        "title": "Car Carrier Transport Optimization System",
        "subtitle": "ระบบคำนวณและวางแผนจัดกลุ่มรถขนส่งสินค้าอัตโนมัติสำหรับฟลีตขนส่งรถยนต์",
        "login_header": "🔓 เข้าสู่ระบบ (System Login)",
        "login_caption": "กรุณากรอก Username และ Password เพื่อเข้าใช้งานระบบตามสิทธิ์",
        "username": "👤 Username (ชื่อผู้ใช้งาน)",
        "password": "🔑 Password (รหัสผ่าน)",
        "login_btn": "🔑 เข้าสู่ระบบ (Sign In)",
        "login_err": "❌ Username หรือ Password ไม่ถูกต้อง!",
        "menu_dashboard": "📊 แดชบอร์ดสรุปภาพรวม",
        "menu_grouping": "🚀 วางแผนจัดกลุ่ม",
        "menu_master": "📂 ข้อมูลมาสเตอร์",
        "menu_cond": "📋 เงื่อนไขการจัดกลุ่ม",
        "menu_fleet": "🚛 ตั้งค่าโควตากองรถ",
        "menu_history": "📜 ประวัติจัดกลุ่มย้อนหลัง",
        "menu_revise": "✏️ แก้ไขและยกเลิกกลุ่ม",
        "main_sub": "🚀 วางแผนและประมวลผลจัดกลุ่มอัตโนมัติ (Main Workspace)",
        "upload_fis_title": "📁 อัปโหลดไฟล์ FIS Ready to Grouping (.xlsx)",
        "upload_fis_desc": "อัปโหลดไฟล์รายการคิวรถที่ต้องการนำมาจัดกลุ่มส่งมอบ",
        "upload_fis_label": "📁 เลือกไฟล์ FIS Ready to Grouping (.xlsx)",
        "process_btn": "🚀 เริ่มคำวณจัดกลุ่มอัตโนมัติ (Process Grouping)",
        "download_btn": "📥 ดาวน์โหลดผลลัพธ์จัดกลุ่ม (.xlsx)",
        "guide_text": "💡 คำแนะนำ: กรุณาเลือกไฟล์ Grouping order (FIS Ready to Grouping) ด้านบนเพื่อกดปุ่มประมวลผล",
    },
    "ENG": {
        "title": "Car Carrier Transport Optimization System",
        "subtitle": "Automated Fleet Grouping & Logistics Optimization Platform for Car Carriers",
        "login_header": "🔓 System Login",
        "login_caption": "Please enter Username and Password to access your assigned role",
        "username": "👤 Username",
        "password": "🔑 Password",
        "login_btn": "🔑 Sign In",
        "login_err": "❌ Invalid Username or Password!",
        "menu_dashboard": "📊 Executive Dashboard",
        "menu_grouping": "🚀 Transport Grouping",
        "menu_master": "📂 Master List",
        "menu_cond": "📋 Grouping Conditions",
        "menu_fleet": "🚛 Fleet Capacity Settings",
        "menu_history": "📜 Execution History",
        "menu_revise": "✏️ Revise & Cancel Grouping",
        "main_sub": "🚀 Transport Grouping Workspace",
        "upload_fis_title": "📁 Upload FIS Ready to Grouping (.xlsx)",
        "upload_fis_desc": "Upload pending car shipment list to process auto grouping",
        "upload_fis_label": "📁 Select FIS Ready to Grouping (.xlsx)",
        "process_btn": "🚀 Process Auto Grouping",
        "download_btn": "📥 Download Result Grouping (.xlsx)",
        "guide_text": "💡 Instruction: Please upload the Grouping order (FIS Ready to Grouping) file above to begin processing.",
    },
}

MODEL_WEIGHT_MASTER = {
    "DOLPHIN": 1615,
    "ATTO 3": 1750,
    "SEAL": 2050,
    "SEAL U": 2020,
    "DENZA D9": 2690,
    "SEAGULL": 1160,
    "M6": 2000,
    "SEALION 5": 1800,
    "SEAL 5": 1600,
}

DEALER_REGION_MAP = {
    "EV-D Ubon Co., Ltd.  (Ubon Ratchathani)": "Northeast",
    "Jinlong Motors Co., Ltd. (Chaengwattana)": "BKK",
    "Metromobile Co., Ltd. (Talingchan)": "BKK",
    "Metromobile Co., Ltd. (Onnut)": "BKK",
    "BKK Automobile Co., Ltd. (Minburi-Ramindra)": "BKK",
    "BKK EV Car Co., Ltd. (Donmuang)": "BKK",
    "Jinlong S-Nakarin Co., Ltd. (Srinagarindra)": "BKK",
    "Autopia Co., Ltd. (Theparak)": "BKK",
    "Yonpiboon Automobile Co., Ltd. (Khonkaen)": "Northeast",
    "Metromobile Co., Ltd. (RCA)": "BKK",
    "B1 Ratchaburi Automotive Co., Ltd. (Ratchaburi)": "West",
}

HISTORY_FILE = "grouping_history.json"

def normalize_key(text):
    if pd.isna(text) or text is None:
        return ""
    text_str = str(text).strip().upper()
    return re.sub(r'[^A-Z0-9]', '', text_str)

def load_history():
    if os.path.exists(HISTORY_FILE):
        try:
            with open(HISTORY_FILE, "r", encoding="utf-8") as f:
                data = json.load(f)
                if data and len(data) > 0:
                    return data
        except Exception:
            pass
    return {}

def save_history(history_data):
    try:
        with open(HISTORY_FILE, "w", encoding="utf-8") as f:
            json.dump(history_data, f, ensure_ascii=False, indent=2)
    except Exception as e:
        st.error(f"ไม่สามารถบันทึกประวัติได้: {e}")

def is_car_ready_to_ship(row, hold_col="HOLD", remark_col="Remark"):
    if pd.notna(row.get(hold_col, None)):
        hold_val = str(row[hold_col]).strip()
        if hold_val != "" and hold_val.lower() != "nan":
            return False, f"HOLD: {hold_val}"

    if pd.notna(row.get(remark_col, None)):
        remark_val = str(row[remark_col]).strip().lower()
        unready_keywords = ["hold", "รอ", "ภายหลัง", "ยังไม่ถึงกำหนด", "รอนัด", "ชะลอ"]
        for kw in unready_keywords:
            if kw in remark_val and remark_val != "nan":
                return False, f"Remark: {str(row[remark_col]).strip()}"

    return True, "Ready"

def process_fis_grouping_adapted(file_bytes, grouping_date_obj, target_regions=["BKK", "Northeast", "West"]):
    grouping_date_str = grouping_date_obj.strftime("%y%m%d")
    grouping_date_display = grouping_date_obj.strftime("%d %b %y")

    df = pd.read_excel(file_bytes)

    pickup_col = "Location" if "Location" in df.columns else "Pick up Location"
    delivery_col = "Delivery Location"
    region_col = "Region"
    model_col = "Model" if "Model" in df.columns else "MODEL NAME"
    group_no_col = "Grouping number"
    group_date_col = "Grouping Date"
    hold_col = "HOLD"
    remark_col = "Remark"
    alloc_date_col = "Allocation Date"

    df["Mapped_Region"] = df[delivery_col].astype(str).str.strip().map(DEALER_REGION_MAP)
    df[region_col] = df[region_col].fillna(df["Mapped_Region"])

    df[group_no_col] = df[group_no_col].astype(object)
    if group_date_col in df.columns:
        df[group_date_col] = df[group_date_col].astype(object)

    df["Ready_Tuple"] = df.apply(lambda r: is_car_ready_to_ship(r, hold_col, remark_col), axis=1)
    df["Ready_Flag"] = df["Ready_Tuple"].apply(lambda x: x[0])
    df["Unready_Reason"] = df["Ready_Tuple"].apply(lambda x: x[1])

    df["Is_Express"] = df[remark_col].astype(str).str.contains("จัดส่งด่วน|ด่วน|express", case=False, na=False)
    if alloc_date_col in df.columns:
        df["_sort_date"] = pd.to_datetime(df[alloc_date_col], errors="coerce")
    else:
        df["_sort_date"] = pd.Timestamp.max

    ready_df = df[(df["Ready_Flag"] == True) & (df[region_col].isin(target_regions))].copy()
    ready_df = ready_df.sort_values(by=["Is_Express", "_sort_date"], ascending=[False, True])

    ready_df["Estimated_Weight_KG"] = (
        ready_df[model_col]
        .astype(str)
        .str.upper()
        .map(lambda x: MODEL_WEIGHT_MASTER.get(x, 1800))
    )

    df["Calc_Group_No"] = ""
    df["Calc_Group_Date"] = ""
    summary_list = []
    prefix = f"SJWD{grouping_date_str}-"
    group_counter = 1

    bkk_ready = ready_df[ready_df[region_col] == "BKK"].copy()
    bkk_dealer_counts = bkk_ready[delivery_col].value_counts()

    for dealer, count in bkk_dealer_counts.items():
        if count >= 5:
            dealer_indices = bkk_ready[bkk_ready[delivery_col] == dealer].index.tolist()
            load_size = min(count, 7)
            group_indices = dealer_indices[:load_size]

            current_group_id = f"{prefix}{group_counter:03d}"
            df.loc[group_indices, "Calc_Group_No"] = current_group_id
            df.loc[group_indices, "Calc_Group_Date"] = grouping_date_display

            group_weight = ready_df.loc[group_indices, "Estimated_Weight_KG"].sum()
            vins_in_group = ready_df.loc[group_indices, "Vin"].astype(str).tolist() if "Vin" in ready_df.columns else []

            summary_list.append(
                {
                    "Grouping ID": current_group_id,
                    "Type": f"Single-Dealer ({len(group_indices)} Load)",
                    "Region": "BKK",
                    "Pick up Locations": ", ".join(map(str, set(ready_df.loc[group_indices, pickup_col]))),
                    "Delivery Locations": str(dealer),
                    "Car Count": len(group_indices),
                    "Total Weight (kg)": group_weight,
                    "VINs": vins_in_group,
                    "Indices": [int(x) for x in group_indices],
                }
            )
            group_counter += 1
            ready_df = ready_df.drop(index=group_indices)

    bkk_rem = ready_df[ready_df[region_col] == "BKK"].index.tolist()
    if len(bkk_rem) >= 6:
        group_size = min(len(bkk_rem), 8)
        group_indices = bkk_rem[:group_size]

        current_group_id = f"{prefix}{group_counter:03d}"
        df.loc[group_indices, "Calc_Group_No"] = current_group_id
        df.loc[group_indices, "Calc_Group_Date"] = grouping_date_display

        group_weight = ready_df.loc[group_indices, "Estimated_Weight_KG"].sum()
        vins_in_group = ready_df.loc[group_indices, "Vin"].astype(str).tolist() if "Vin" in ready_df.columns else []

        summary_list.append(
            {
                "Grouping ID": current_group_id,
                "Type": f"Trailer ({len(group_indices)} Load)",
                "Region": "BKK",
                "Pick up Locations": ", ".join(map(str, set(ready_df.loc[group_indices, pickup_col]))),
                "Delivery Locations": ", ".join(map(str, set(ready_df.loc[group_indices, delivery_col]))),
                "Car Count": len(group_indices),
                "Total Weight (kg)": group_weight,
                "VINs": vins_in_group,
                "Indices": [int(x) for x in group_indices],
            }
        )
        group_counter += 1
        ready_df = ready_df.drop(index=group_indices)

    for reg in ["Northeast", "West"]:
        reg_indices = ready_df[ready_df[region_col] == reg].index.tolist()
        while len(reg_indices) >= 6:
            target_count = 8 if len(reg_indices) >= 8 else (7 if len(reg_indices) >= 7 else 6)
            group_indices = reg_indices[:target_count]

            current_group_id = f"{prefix}{group_counter:03d}"
            df.loc[group_indices, "Calc_Group_No"] = current_group_id
            df.loc[group_indices, "Calc_Group_Date"] = grouping_date_display

            group_weight = ready_df.loc[group_indices, "Estimated_Weight_KG"].sum()
            vins_in_group = ready_df.loc[group_indices, "Vin"].astype(str).tolist() if "Vin" in ready_df.columns else []

            summary_list.append(
                {
                    "Grouping ID": current_group_id,
                    "Type": f"Trailer ({len(group_indices)} Load)",
                    "Region": reg,
                    "Pick up Locations": ", ".join(map(str, set(ready_df.loc[group_indices, pickup_col]))),
                    "Delivery Locations": ", ".join(map(str, set(ready_df.loc[group_indices, delivery_col]))),
                    "Car Count": len(group_indices),
                    "Total Weight (kg)": group_weight,
                    "VINs": vins_in_group,
                    "Indices": [int(x) for x in group_indices],
                }
            )
            group_counter += 1
            ready_df = ready_df.drop(index=group_indices)
            reg_indices = ready_df[ready_df[region_col] == reg].index.tolist()

    wb = openpyxl.load_workbook(file_bytes)
    ws = wb.active

    headers = [cell.value for cell in ws[1]]
    g_no_col_idx = headers.index(group_no_col) + 1
    region_col_idx = headers.index(region_col) + 1

    for idx, row in df.iterrows():
        excel_row_num = idx + 2
        calc_no = row["Calc_Group_No"]
        calc_date = row["Calc_Group_Date"]
        region_val = row[region_col]

        ws.cell(row=excel_row_num, column=region_col_idx, value=region_val)

        if calc_no != "":
            ws.cell(row=excel_row_num, column=g_no_col_idx, value=calc_no)
            if group_date_col in headers:
                g_date_col_idx = headers.index(group_date_col) + 1
                ws.cell(row=excel_row_num, column=g_date_col_idx, value=calc_date)

    output_buffer = io.BytesIO()
    wb.save(output_buffer)
    output_buffer.seek(0)

    total_cars = len(df)
    return output_buffer, pd.DataFrame(summary_list), total_cars, df

# --- STREAMLIT CONFIG ---
st.set_page_config(
    page_title="SIAM JWD LOGISTICS - Car Carrier TMS",
    page_icon="🚚",
    layout="wide",
    initial_sidebar_state="expanded",
)

TIMEOUT_SECONDS = 600
if "last_activity" in st.session_state and st.session_state.get("authenticated", False):
    if time.time() - st.session_state["last_activity"] > TIMEOUT_SECONDS:
        st.session_state["authenticated"] = False
        st.session_state["user_info"] = None
        st.warning("⏱️ หมดเวลาการใช้งานระบบเนื่องจากไม่มีการเคลื่อนไหวเกิน 10 นาที กรุณาเข้าสู่ระบบใหม่")
        st.rerun()

st.session_state["last_activity"] = time.time()

if "lang" not in st.session_state:
    st.session_state["lang"] = "TH"

car_carrier_bg_url = "https://images.unsplash.com/photo-1601584115197-04ecc0da31d7?auto=format&fit=crop&w=1920&q=80"

# --- SIMPLE CSS & THEME ---
st.markdown(
    f\"\"\"
    <style>
    .block-container {{
        padding-top: 1rem !important;
        padding-bottom: 1rem !important;
        padding-left: 1.5rem !important;
        padding-right: 1.5rem !important;
        max-width: 100% !important;
    }}
    [data-testid="stHeader"] {{ background: transparent !important; }}
    footer {{ visibility: hidden !important; height: 0px !important; }}
    a.anchor-link {{ display: none !important; }}
    
    [data-testid="stSidebar"] {{
        background-color: #f8fafc !important;
        border-right: 1px solid #e2e8f0 !important;
        width: 250px !important;
    }}
    [data-testid="stSidebar"] * {{ color: #1e293b !important; }}
    [data-testid="stSidebar"] label {{
        padding: 8px 12px !important;
        border-radius: 8px !important;
        font-size: 13px !important;
        font-weight: 500 !important;
        cursor: pointer !important;
    }}
    
    .login-bg {{
        background: linear-gradient(rgba(11, 37, 69, 0.85), rgba(11, 37, 69, 0.90)), url('{car_carrier_bg_url}');
        background-size: cover;
        background-position: center;
        padding: 22px 20px;
        border-radius: 12px;
        color: white;
        margin-bottom: 15px;
    }}
    
    .clean-card {{
        background-color: #ffffff;
        border-left: 5px solid #0066B3;
        border-radius: 8px;
        padding: 14px;
        box-shadow: 0 2px 8px rgba(0,0,0,0.06);
        margin-bottom: 10px;
    }}
    .clean-card-red {{
        background-color: #ffffff;
        border-left: 5px solid #ED1C24;
        border-radius: 8px;
        padding: 14px;
        box-shadow: 0 2px 8px rgba(0,0,0,0.06);
        margin-bottom: 10px;
    }}
    .sidebar-user-box {{
        background: #ffffff;
        border: 1px solid #cbd5e1;
        border-radius: 8px;
        padding: 8px 12px;
        margin-bottom: 12px;
    }}
    </style>
    \"\"\",
    unsafe_allow_html=True,
)

# --- LOGIN SYSTEM ---
if "authenticated" not in st.session_state:
    st.session_state["authenticated"] = False
    st.session_state["user_info"] = None

if not st.session_state["authenticated"]:
    st.write("")
    l_col1, l_col2 = st.columns([0.80, 0.20])
    with l_col2:
        selected_lang = st.radio(
            "Language",
            ["TH", "ENG"],
            index=0 if st.session_state["lang"] == "TH" else 1,
            horizontal=True,
            label_visibility="collapsed",
            key="login_lang_radio"
        )
        if selected_lang != st.session_state["lang"]:
            st.session_state["lang"] = selected_lang
            st.rerun()

    txt = T[st.session_state["lang"]]

    st.markdown(
        f\"\"\"
        <div class="login-bg">
            <div style="text-align:center;">
                <span style="color:#ED1C24; font-size:42px; font-weight:900;">SIAM </span>
                <span style="color:#ffffff; font-size:42px; font-weight:900;">JWD</span><br>
                <span style="color:#cbd5e1; font-size:12px; letter-spacing:5px; font-weight:bold;">LOGISTICS</span>
                <h3 style="color:#ffffff; margin-top:8px; margin-bottom:4px; font-weight:800;">{txt['title']}</h3>
                <p style="color:#e2e8f0; font-size:13px; margin-bottom:0;">{txt['subtitle']}</p>
            </div>
        </div>
        \"\"\",
        unsafe_allow_html=True,
    )

    col_l1, col_l2, col_l3 = st.columns([1, 1.5, 1])
    with col_l2:
        st.markdown(f"#### **{txt['login_header']}**")
        st.caption(txt["login_caption"])

        with st.form("login_form"):
            username_input = st.text_input(txt["username"])
            password_input = st.text_input(txt["password"], type="password")
            submit_login = st.form_submit_button(txt["login_btn"], use_container_width=True)

            if submit_login:
                user = USER_DB.get(username_input.strip().lower())
                if user and user["password"] == password_input:
                    st.session_state["authenticated"] = True
                    st.session_state["last_activity"] = time.time()
                    st.session_state["user_info"] = {
                        "username": username_input,
                        "name": user["name"],
                        "role": user["role"],
                    }
                    st.success("Success!")
                    st.rerun()
                else:
                    st.error(txt["login_err"])
    st.stop()

txt = T[st.session_state["lang"]]
current_user = st.session_state["user_info"]

st.sidebar.markdown(
    f\"\"\"
    <div class="sidebar-user-box">
        <div style="font-weight:800; font-size:13px; color:#0b2545;">👤 {current_user['name']}</div>
        <div style="font-size:11px; color:#64748b;">🔑 {current_user['role']}</div>
    </div>
    \"\"\",
    unsafe_allow_html=True,
)

menu_options = [
    txt["menu_dashboard"],
    txt["menu_grouping"],
    txt["menu_master"],
    txt["menu_cond"],
    txt["menu_fleet"],
    txt["menu_history"],
    txt["menu_revise"],
]

active_feature = st.sidebar.radio("Navigation", menu_options, index=0, label_visibility="collapsed")

st.sidebar.markdown("---")
st.sidebar.caption("SIAM JWD LOGISTICS CO., LTD.")

head_col1, head_col2 = st.columns([0.70, 0.30])
with head_col1:
    st.markdown("### **Car Carrier Transport Optimization System**")
    st.caption(txt["subtitle"])

with head_col2:
    if st.button("🚪 ออกจากระบบ", key="btn_logout_main", use_container_width=True):
        st.session_state["authenticated"] = False
        st.session_state["user_info"] = None
        st.rerun()

st.divider()

# --- 0. DASHBOARD ---
if active_feature == txt["menu_dashboard"]:
    st.subheader("📊 แดชบอร์ดสรุปภาพรวมแผนจัดกลุ่มรถขนส่ง (Executive Fleet Dashboard)")
    history_data = load_history()

    if not history_data:
        st.info("💡 ยังไม่มีข้อมูลการจัดกลุ่มในระบบ")
    else:
        available_records = sorted(list(history_data.keys()), reverse=True)
        selected_record_key = st.selectbox("📅 เลือกประวัติรอบการจัดกลุ่ม:", available_records)

        if selected_record_key in history_data:
            rec = history_data[selected_record_key]
            if "full_details" in rec and rec["full_details"]:
                df_dash = pd.DataFrame(rec["full_details"])
                pickup_col = "Location" if "Location" in df_dash.columns else "Pick up Location"
                group_col = "Grouping number" if "Grouping number" in df_dash.columns else "Calc_Group_No"

                df_dash["Ready_Tuple"] = df_dash.apply(lambda r: is_car_ready_to_ship(r), axis=1)
                df_dash["Ready_Flag"] = df_dash["Ready_Tuple"].apply(lambda x: x[0])
                df_dash["Unready_Reason"] = df_dash["Ready_Tuple"].apply(lambda x: x[1])

                if group_col in df_dash.columns:
                    df_dash["Is_Grouped"] = df_dash[group_col].notna() & (df_dash[group_col].astype(str).str.strip() != "") & (df_dash[group_col].astype(str).str.strip() != "เศษรอ Mix") & (df_dash[group_col].astype(str).str.strip() != "nan")
                else:
                    df_dash["Is_Grouped"] = False

                total_cars = len(df_dash)
                grouped_cars = df_dash["Is_Grouped"].sum()
                ungrouped_cars = total_cars - grouped_cars
                unready_cars = total_cars - df_dash["Ready_Flag"].sum()

                d1, d2, d3, d4 = st.columns(4)
                d1.metric("จำนวนรถทั้งหมด", f"{total_cars} คัน")
                d2.metric("จัดกลุ่มสำเร็จแล้ว", f"{grouped_cars} คัน")
                d3.metric("คงเหลือยังไม่ได้จัดกลุ่ม", f"{ungrouped_cars} คัน")
                d4.metric("รถติด Hold/Unready", f"{unready_cars} คัน")

                st.divider()
                col_db1, col_db2 = st.columns(2)
                with col_db1:
                    st.markdown("### **ยอดรถคงเหลือแยกตามยาร์ด**")
                    if pickup_col in df_dash.columns:
                        yard_summary = df_dash.groupby(pickup_col).agg(
                            Total_Cars=("Vin", "count"),
                            Grouped=("Is_Grouped", "sum"),
                            Remaining_Ungrouped=("Is_Grouped", lambda x: (~x).sum()),
                        ).reset_index()
                        st.dataframe(yard_summary, use_container_width=True)

                with col_db2:
                    st.markdown("### **จำแนกหมวดหมู่รถติด Hold**")
                    unready_df = df_dash[df_dash["Ready_Flag"] == False]
                    if not unready_df.empty:
                        st.dataframe(unready_df["Unready_Reason"].value_counts().reset_index(), use_container_width=True)

# --- 1. GROUPING WORKSPACE ---
elif active_feature == txt["menu_grouping"]:
    st.subheader("🚀 วางแผนจัดกลุ่ม (Transport Grouping)")
    tab_auto, tab_manual = st.tabs(["🚀 จัดกลุ่ม (Auto grouping)", "📥 จัดกลุ่ม Manual"])

    with tab_auto:
        uploaded_file = st.file_uploader(txt["upload_fis_label"], type=["xlsx", "xls"], key="main_fis_auto")
        if uploaded_file:
            selected_regions = st.multiselect("📍 เลือกภูมิภาคที่ต้องการจัดกลุ่ม:", ["BKK", "Northeast", "West", "North", "East", "Central", "South"], default=["BKK", "Northeast", "West"])
            if st.button(txt["process_btn"], type="primary", use_container_width=True):
                out_buffer, df_summary, total_cars, df_processed = process_fis_grouping_adapted(io.BytesIO(uploaded_file.getvalue()), datetime.now(), target_regions=selected_regions)
                st.dataframe(df_summary, use_container_width=True)
                
                history = load_history()
                date_key = datetime.now().strftime("%Y-%m-%d_%H%M%S_Auto")
                history[date_key] = {
                    "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    "mode": "Auto Grouping",
                    "total_cars": total_cars,
                    "grouped_cars": int(df_summary["Car Count"].sum()) if not df_summary.empty else 0,
                    "total_groups": len(df_summary),
                    "summary": df_summary.to_dict(orient="records"),
                    "full_details": df_processed.fillna("").astype(str).to_dict(orient="records"),
                }
                save_history(history)
                st.download_button(label=txt["download_btn"], data=out_buffer, file_name=f"FIS_Grouped_{datetime.now().strftime('%Y%m%d')}.xlsx")

    with tab_manual:
        actual_file = st.file_uploader("เลือกไฟล์ FIS ที่จัดกลุ่ม Manual แล้ว (.xlsx)", type=["xlsx", "xls"], key="manual_import_file")
        if actual_file:
            xls_m = pd.ExcelFile(actual_file)
            all_rows = []
            if "Sheet3  (2)" in xls_m.sheet_names:
                df_main = pd.read_excel(actual_file, sheet_name="Sheet3  (2)")
                vin_map = {}
                for sname in xls_m.sheet_names:
                    if sname != "Sheet3  (2)" and "Sheet" not in sname:
                        df_s = pd.read_excel(actual_file, sheet_name=sname)
                        if len(df_s) > 1:
                            df_s.columns = df_s.iloc[0].values
                            data_s = df_s.iloc[1:].copy()
                            data_s['Groupping  Number'] = data_s['Groupping  Number'].ffill()
                            for _, r_s in data_s.iterrows():
                                v_s = str(r_s.get('Vin', '')).strip()
                                g_s = str(r_s.get('Groupping  Number', '')).strip()
                                if v_s and v_s != 'nan':
                                    vin_map[v_s] = g_s
                
                df_main['Grouping number'] = df_main['Vin'].astype(str).str.strip().map(vin_map).fillna(df_main['Grouping number'])
                df_act = df_main.copy()
            else:
                df_act = pd.read_excel(actual_file)

            group_col = "Grouping number" if "Grouping number" in df_act.columns else None
            if group_col and group_col in df_act.columns:
                df_act[group_col] = df_act[group_col].fillna("เศษรอ Mix")
                act_grouped = df_act[df_act[group_col].notna() & (~df_act[group_col].astype(str).str.strip().isin(["เศษรอ Mix", "", "nan"]))].copy()
                st.success(f"✅ อ่านข้อมูลสำเร็จทั้งหมด {len(df_act)} คัน (จัดกลุ่มแล้ว {len(act_grouped)} คัน)")

                if st.button("💾 บันทึกเข้า History Benchmark", type="primary", use_container_width=True):
                    history = load_history()
                    date_key = datetime.now().strftime("%Y-%m-%d_%H%M%S_Manual")
                    df_act_clean = df_act.copy().astype(str)
                    
                    sum_df = act_grouped.groupby(group_col).agg(
                        Car_Count=("Vin", "count"),
                        Region=("Region", lambda x: ", ".join(map(str, set(x)))),
                        Delivery_Locations=("Delivery Location", lambda x: ", ".join(map(str, set(x)))),
                    ).reset_index()
                    sum_df.columns = ["Grouping ID", "Car Count", "Region", "Delivery Locations"]

                    history[date_key] = {
                        "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                        "mode": "Manual Actual Import",
                        "total_cars": len(df_act),
                        "grouped_cars": len(act_grouped),
                        "total_groups": len(sum_df),
                        "summary": sum_df.to_dict(orient="records"),
                        "full_details": df_act_clean.fillna("").to_dict(orient="records"),
                    }
                    save_history(history)
                    st.balloons()
                    st.success("🎉 บันทึกข้อมูลคิวรถทั้งหมดสำเร็จเรียบร้อย!")

# --- 2. MASTER LIST ---
elif active_feature == txt["menu_master"]:
    st.subheader(f"📂 {txt['menu_master']}")
    if current_user["role"] in ["Admin", "Project Manager"]:
        master_up = st.file_uploader("📂 Upload Master Dealer (Region).xlsx:", type=["xlsx", "xls"], key="menu_master_file")
        if master_up:
            df_master_view = pd.read_excel(master_up)
            st.session_state["master_df_stored"] = df_master_view
            st.dataframe(df_master_view, use_container_width=True)
        elif "master_df_stored" in st.session_state:
            st.dataframe(st.session_state["master_df_stored"], use_container_width=True)

# --- 3. CONDITIONS ---
elif active_feature == txt["menu_cond"]:
    st.subheader(f"📋 {txt['menu_cond']}")
    st.info("เงื่อนไข: เน้นจัดกลุ่มส่งมอบดีลเลอร์เดียวก่อน (5-7 คัน) ในเขตกรุงเทพฯ และเรียงคิวรถด่วนขึ้นก่อน")

# --- 4. FLEET ---
elif active_feature == txt["menu_fleet"]:
    st.subheader(f"🚛 {txt['menu_fleet']}")
    st.success("💾 โควตากองรถอยู่ในสถานะพร้อมใช้งาน")

# --- 5. HISTORY (UNIVERSAL ADVANCED SEARCH) ---
elif active_feature == txt["menu_history"]:
    st.subheader("📜 ประวัติจัดกลุ่มย้อนหลังและการค้นหาคันรถ")
    history_data = load_history()
    
    if not history_data:
        st.info("💡 ยังไม่มีข้อมูลในประวัติระบบ")
    else:
        all_cars_list = []
        for hkey, hrec in history_data.items():
            if "full_details" in hrec and hrec["full_details"]:
                for row in hrec["full_details"]:
                    r_copy = dict(row)
                    r_copy["History_Run"] = hkey
                    all_cars_list.append(r_copy)

        df_all_history = pd.DataFrame(all_cars_list)

        st.markdown(
            \"\"\"
            <div class="clean-card">
                <h4 style="color:#0066B3; margin:0;">🔍 ค้นหาข้อมูลแบบรวม (ค้นหาได้จาก Grouping number, VIN, วันที่ หรือสถานที่ส่ง)</h4>
            </div>
            \"\"\",
            unsafe_allow_html=True,
        )

        c_srch1, c_srch2 = st.columns([0.80, 0.20])
        with c_srch1:
            univ_search = st.text_input(
                "🔎 พิมพ์คำค้นหา (VIN, Grouping number, วันที่ Allocation, ดีลเลอร์ ฯลฯ):",
                placeholder="พิมพ์เพื่อค้นหา เช่น ATL260821-001 หรือ SJWD260821-006 หรือ LGXCE4...",
                label_visibility="collapsed"
            )
        with c_srch2:
            if st.button("❌ ล้างการค้นหา", use_container_width=True):
                st.rerun()

        if univ_search.strip():
            query_str = univ_search.strip().lower()
            mask = df_all_history.astype(str).apply(lambda col: col.str.lower().str.contains(query_str, na=False)).any(axis=1)
            filtered_df = df_all_history[mask].copy()
            st.success(f"✅ พบข้อมูลทั้งหมด {len(filtered_df)} รายการที่ตรงกับคำค้นหา '{univ_search}'")
        else:
            filtered_df = df_all_history.copy()
            st.caption(f"แสดงรายการรถทั้งหมดในระบบประวัติ ({len(filtered_df)} รายการ)")

        excel_target_cols = [
            "Dealer Code", "Name", "Vin", "MODEL NAME", "Model",
            "Color", "Location", "Delivery Location", "Allocation Date", "Grouping number"
        ]
        
        valid_cols = [c for c in excel_target_cols if c in filtered_df.columns]
        if not valid_cols:
            valid_cols = filtered_df.columns.tolist()

        st.dataframe(filtered_df[valid_cols], use_container_width=True, hide_index=True)

# --- 6. REVISE & SEARCH MODULE (REQUIRE TYPING FIRST + EXACT EXCEL TABLE COLUMNS) ---
elif active_feature == txt["menu_revise"]:
    st.subheader("✏️ แก้ไขและยกเลิกกลุ่ม (Revise Grouping Number)")

    history_data = load_history()

    all_groups_pool = set()
    for hkey, hrec in history_data.items():
        if "full_details" in hrec and hrec["full_details"]:
            for r_item in hrec["full_details"]:
                gid = str(r_item.get("Grouping number", r_item.get("Calc_Group_No", r_item.get("Grouping ID", "")))).strip()
                if gid and gid not in ["nan", "เศษรอ Mix", ""]:
                    all_groups_pool.add(gid)

    sorted_groups = sorted(list(all_groups_pool))

    st.markdown(
        \"\"\"
        <div class="clean-card">
            <h3 style="color:#0066B3; margin:0;">🔍 ค้นหา Grouping Number ที่ต้องการแก้ไข / ยกเลิก</h3>
        </div>
        \"\"\",
        unsafe_allow_html=True,
    )

    if "confirmed_group_id" not in st.session_state:
        st.session_state["confirmed_group_id"] = ""

    col_input, col_confirm, col_clear = st.columns([0.70, 0.15, 0.15])
    
    with col_input:
        typed_input = st.text_input(
            "🔎 พิมพ์ Group number ที่ต้องการค้นหา:",
            placeholder="กรอก Group number เช่น SJWD260821-006 หรือ ATL...",
            label_visibility="collapsed",
            key="input_revise_text_field"
        )
    
    matched_suggestions = []
    if typed_input.strip():
        norm_typed = normalize_key(typed_input)
        matched_suggestions = [g for g in sorted_groups if norm_typed in normalize_key(g)]

    if matched_suggestions:
        selected_from_sug = st.selectbox("💡 รายการที่แนะนำตรงกับคำค้นหา:", options=matched_suggestions, index=0)
    else:
        selected_from_sug = typed_input.strip()

    with col_confirm:
        if st.button("🔍 ค้นหา", type="primary", use_container_width=True):
            if selected_from_sug.strip():
                st.session_state["confirmed_group_id"] = selected_from_sug.strip()
            else:
                st.warning("กรุณาพิมพ์ Group number ก่อนกดค้นหาครับ")

    with col_clear:
        if st.button("❌ ล้างค่า", use_container_width=True):
            st.session_state["confirmed_group_id"] = ""
            st.rerun()

    target_group_id = st.session_state["confirmed_group_id"]

    if not target_group_id:
        st.info("💡 พิมพ์ Group number ในช่องด้านบน แล้วกดปุ่ม **'🔍 ค้นหา'** เพื่อเริ่มดูและแก้ไขคิวรถในกลุ่ม")
    else:
        st.divider()
        st.markdown(f"### **📋 ผลการค้นหาสำหรับ Grouping ID: `{target_group_id}`**")

        norm_target = normalize_key(target_group_id)
        matched_records = []

        for hkey, hrec in history_data.items():
            if "full_details" in hrec and hrec["full_details"]:
                df_temp = pd.DataFrame(hrec["full_details"])
                for gcol_t in ["Grouping number", "Calc_Group_No", "Grouping ID"]:
                    if gcol_t in df_temp.columns:
                        normalized_series = df_temp[gcol_t].apply(normalize_key)
                        mask = normalized_series.str.contains(norm_target, na=False)
                        if mask.any():
                            matched_records.append((hkey, hrec, df_temp, gcol_t, mask))
                            break

        if not matched_records:
            st.error(f"❌ ไม่พบข้อมูลสำหรับ Grouping Number: `{target_group_id}`")
        else:
            hkey, hrec, df_matched, gcol, mask_match = matched_records[0]
            group_vins_df = df_matched[mask_match].copy().reset_index(drop=True)

            st.success(f"✅ พบรถในกลุ่มนี้ทั้งหมด {len(group_vins_df)} คัน")
            
            c_all1, c_all2 = st.columns([0.25, 0.75])
            with c_all1:
                select_all_hdr = st.checkbox("☑️ เลือกทั้งหมด (Select All)", value=False, key=f"chk_all_hdr_{target_group_id}")

            target_excel_cols = [
                "Dealer Code", "Name", "Vin", "MODEL NAME", "Model",
                "Color", "Location", "Delivery Location", "Allocation Date", "Grouping number"
            ]
            
            display_cols = [c for c in target_excel_cols if c in group_vins_df.columns]
            if not display_cols:
                display_cols = group_vins_df.columns.tolist()

            table_data = group_vins_df[display_cols].copy()
            table_data.insert(0, "เลือกถอดออก", select_all_hdr)

            edited_table = st.data_editor(
                table_data,
                column_config={
                    "เลือกถอดออก": st.column_config.CheckboxColumn("เลือกถอดออก", default=select_all_hdr),
                    "Dealer Code": st.column_config.TextColumn("Dealer Code", disabled=True),
                    "Name": st.column_config.TextColumn("Name", disabled=True),
                    "Vin": st.column_config.TextColumn("Vin", disabled=True),
                    "MODEL NAME": st.column_config.TextColumn("MODEL NAME", disabled=True),
                    "Model": st.column_config.TextColumn("Model", disabled=True),
                    "Color": st.column_config.TextColumn("Color", disabled=True),
                    "Location": st.column_config.TextColumn("Location", disabled=True),
                    "Delivery Location": st.column_config.TextColumn("Delivery Location", disabled=True),
                    "Allocation Date": st.column_config.TextColumn("Allocation Date", disabled=True),
                    "Grouping number": st.column_config.TextColumn("Grouping number", disabled=True),
                },
                hide_index=True,
                use_container_width=True,
                key=f"editor_group_{target_group_id}"
            )

            selected_vins_to_remove = edited_table[edited_table["เลือกถอดออก"] == True]["Vin"].tolist()

            st.write("")
            col_act1, col_act2 = st.columns(2)

            with col_act1:
                if st.button(f"🚨 ยกเลิกกลุ่ม {target_group_id} ทั้งหมด ({len(group_vins_df)} คัน)", type="primary", use_container_width=True):
                    df_matched.loc[mask_match, gcol] = "เศษรอ Mix"
                    history_data[hkey]["full_details"] = df_matched.fillna("").astype(str).to_dict(orient="records")
                    save_history(history_data)
                    st.balloons()
                    st.success("🎉 ยกเลิกกลุ่มเรียบร้อยแล้ว!")
                    time.sleep(1)
                    st.session_state["confirmed_group_id"] = ""
                    st.rerun()

            with col_act2:
                if st.button(f"❌ ถอดเฉพาะคันที่เลือก ({len(selected_vins_to_remove)} คัน)", use_container_width=True):
                    if selected_vins_to_remove:
                        df_matched.loc[df_matched["Vin"].isin(selected_vins_to_remove), gcol] = "เศษรอ Mix"
                        history_data[hkey]["full_details"] = df_matched.fillna("").astype(str).to_dict(orient="records")
                        save_history(history_data)
                        st.success(f"ถอด VIN จำนวน {len(selected_vins_to_remove)} คันสำเร็จ!")
                        time.sleep(1)
                        st.rerun()
                    else:
                        st.warning("⚠️ กรุณาติ๊กเลือกช่อง 'เลือกถอดออก' ในตารางด้านบนก่อนครับ")
"""
